"""Test-case matrix workbook in the reference ``TC-order-workflow.xlsx`` layout.

Sheet **Test Cases**: columns ``TC ID | Title | Preconditions | Steps | Expected
Result | Type | Automated | Linked Story/Req | Notes`` — header row Arial 10
bold white on ``2F5496`` (wrapped, centred, 30 pt high), frozen at ``A2``, body
rows Arial 10 on ``F2F2F2`` with thin borders, top-aligned and wrapped, the
reference column widths.

Sheet **Summary**: a 14 pt ``2F5496`` title, ``Linked TDD:`` / ``Linked Epic:`` /
``Automation:`` rows, *Totals by Automation Status* (Automated (Yes) / Manual /
Planned / Total Test Cases), *Totals by Type* (one row per Type in the reference
vocabulary order, then any extra types) and *Notes*. Totals are written as
literal numbers computed here (the reference uses ``COUNTIF`` formulas; literal
values stay readable by any tool without an Excel recalculation).

:func:`read_test_case_rows` reads a matrix in this layout back (used to merge the
knowledge base's original rows into the Phase 2 preview and by Phase 4).
"""

from __future__ import annotations

import io
from collections.abc import Iterable, Sequence
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, ConfigDict, Field

from .package import stabilise_package

TC_COLUMNS: tuple[str, ...] = (
    "TC ID",
    "Title",
    "Preconditions",
    "Steps",
    "Expected Result",
    "Type",
    "Automated",
    "Linked Story/Req",
    "Notes",
)
_COLUMN_WIDTHS = (8, 30, 32, 30, 48, 20, 15, 16, 30)
#: The reference "Type" vocabulary, in the order the Summary sheet lists it.
TC_TYPES: tuple[str, ...] = (
    "Functional",
    "Functional / Compensation",
    "Functional / Cancellation",
    "Functional / Validation",
    "Reliability / Idempotency",
    "Reliability",
    "Non-functional / Scalability",
    "Non-functional / Observability",
    "Edge case",
)
HEADER_FILL = "2F5496"
_FIXED_CREATED = datetime(2026, 1, 1, tzinfo=UTC)
BODY_FILL = "F2F2F2"


class TestCaseRow(BaseModel):
    """One row of the matrix (columns in :data:`TC_COLUMNS` order)."""

    __test__ = False  # not a pytest class despite the name
    model_config = ConfigDict(extra="ignore")

    tc_id: str = ""
    title: str = ""
    preconditions: str = ""
    steps: str = ""
    expected: str = ""
    type: str = ""
    automated: str = ""
    linked: str = ""
    notes: str = ""

    def values(self) -> list[str]:
        return [
            self.tc_id,
            self.title,
            self.preconditions,
            self.steps,
            self.expected,
            self.type,
            self.automated,
            self.linked,
            self.notes,
        ]

    @classmethod
    def from_values(cls, values: Sequence[object]) -> TestCaseRow:
        cells = ["" if v is None else str(v).strip() for v in values] + [""] * 9
        return cls(
            tc_id=cells[0],
            title=cells[1],
            preconditions=cells[2],
            steps=cells[3],
            expected=cells[4],
            type=cells[5],
            automated=cells[6],
            linked=cells[7],
            notes=cells[8],
        )


class TestCaseSummary(BaseModel):
    """The Summary sheet's free fields; totals are computed from the rows."""

    __test__ = False
    model_config = ConfigDict(extra="ignore")

    title: str = "Test Case Matrix"
    linked_tdd: str = ""
    linked_epic: str = ""
    automation: str = ""
    notes: list[str] = Field(default_factory=list)


def totals_by_automation(rows: Sequence[TestCaseRow]) -> dict[str, int]:
    values = [r.automated.strip().lower() for r in rows]
    return {
        "Automated (Yes)": sum(1 for v in values if v == "yes"),
        "Manual": sum(1 for v in values if v.startswith("manual")),
        "Planned": sum(1 for v in values if v == "planned"),
        "Total Test Cases": len(rows),
    }


def totals_by_type(rows: Sequence[TestCaseRow]) -> list[tuple[str, int]]:
    seen: dict[str, int] = dict.fromkeys(TC_TYPES, 0)
    for row in rows:
        key = row.type.strip()
        if key:
            seen[key] = seen.get(key, 0) + 1
    return list(seen.items())


def write_test_case_matrix(rows: Iterable[TestCaseRow], summary: TestCaseSummary) -> bytes:
    """Serialise the workbook; deterministic for identical input."""
    rows = list(rows)
    wb = Workbook()
    # Fixed core properties keep the bytes deterministic for identical input.
    wb.properties.creator = "workflow-compiler"
    wb.properties.lastModifiedBy = "workflow-compiler"
    wb.properties.created = _FIXED_CREATED
    wb.properties.modified = _FIXED_CREATED
    ws = wb.active
    ws.title = "Test Cases"
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
    body_font = Font(name="Arial", size=10)
    body_fill = PatternFill(fill_type="solid", fgColor=BODY_FILL)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, name in enumerate(TC_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = _COLUMN_WIDTHS[col - 1]
    ws.row_dimensions[1].height = 30
    for r, row in enumerate(rows, start=2):
        for col, value in enumerate(row.values(), start=1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.font = body_font
            cell.fill = body_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
    ws.freeze_panes = "A2"

    ss = wb.create_sheet("Summary")
    ss.column_dimensions["A"].width = 32
    ss.column_dimensions["B"].width = 28
    ss["A1"] = summary.title
    ss["A1"].font = Font(name="Arial", size=14, bold=True, color=HEADER_FILL)
    ss.row_dimensions[1].height = 17.35
    label_font = Font(name="Calibri", size=11, bold=True)
    ss["A3"], ss["B3"] = "Linked TDD:", summary.linked_tdd
    ss["A4"], ss["B4"] = "Linked Epic:", summary.linked_epic
    ss["A5"], ss["B5"] = "Automation:", summary.automation
    for ref in ("A3", "A4", "A5"):
        ss[ref].font = label_font
    row_no = 7
    ss.cell(row=row_no, column=1, value="Totals by Automation Status").font = label_font
    for label, count in totals_by_automation(rows).items():
        row_no += 1
        ss.cell(row=row_no, column=1, value=label)
        ss.cell(row=row_no, column=2, value=count)
    row_no += 2
    ss.cell(row=row_no, column=1, value="Totals by Type").font = label_font
    for label, count in totals_by_type(rows):
        row_no += 1
        ss.cell(row=row_no, column=1, value=label)
        ss.cell(row=row_no, column=2, value=count)
    row_no += 3
    ss.cell(row=row_no, column=1, value="Notes").font = label_font
    for note in summary.notes:
        row_no += 1
        ss.cell(row=row_no, column=1, value=note)

    buffer = io.BytesIO()
    wb.save(buffer)
    return stabilise_package(buffer.getvalue())


def read_test_case_rows(data: bytes) -> list[TestCaseRow]:
    """Rows of the first sheet whose header row starts with ``TC ID`` (else ``[]``)."""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            iterator = ws.iter_rows(values_only=True)
            header: tuple[Any, ...] | None = next(iterator, None)
            if header is None or not header or str(header[0] or "").strip() != TC_COLUMNS[0]:
                continue
            names = [str(h or "").strip() for h in header]
            index = [names.index(c) if c in names else -1 for c in TC_COLUMNS]
            rows: list[TestCaseRow] = []
            for raw in iterator:
                if raw is None or not any(v is not None and str(v).strip() for v in raw):
                    continue
                values = [raw[i] if 0 <= i < len(raw) else None for i in index]
                row = TestCaseRow.from_values(values)
                if row.tc_id:
                    rows.append(row)
            return rows
        return []
    finally:
        wb.close()
